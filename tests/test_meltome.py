from pathlib import Path
import openpyxl
from prosurf.io.meltome import canonical_uniprot, species_datasets, organism_tms


def test_canonical_uniprot_parses_and_filters():
    assert canonical_uniprot("P61626_LYZ") == "P61626"
    assert canonical_uniprot("Q9UNM6-2_PSMD13") == "Q9UNM6"
    assert canonical_uniprot("O50146_lysY") == "O50146"
    assert canonical_uniprot("notanid_xyz") is None


def _make_s1(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meltome data set"
    ws.append(["Dataset ID", "Dataset source", "Organism", "Strain/tissue", "OGT [°C]"])
    ws.append(["ma:0008", "TUM", "Thermus thermophilus", "HB27", 70])
    ws.append(["ma:0009", "TUM", "Thermus thermophilus", "HB27", 70])
    ws.append(["ma:0003", "TUM", "Escherichia coli", "K12", 37])
    ws.append(["ma:0001", "TUM", "Oleispira antarctica", "RB-8", 15])
    wb.save(path)


def _make_s4(path):
    wb = openpyxl.Workbook()
    first = True
    for sheet in ["ma_0008", "ma_0009"]:
        ws = wb.create_sheet(sheet) if not first else wb.active
        if first:
            ws.title = sheet
            first = False
        ws.append(["Protein ID", "Melting point [°C]", "class"])
    # ma_0008: O50147 melts at 80, P0 non-melter (None)
    wb["ma_0008"].append(["O50147_lysZ", 80.0, "medium"])
    wb["ma_0008"].append(["O66271_fumC", None, "non-melter"])
    # ma_0009: O50147 again at 82 -> median across datasets = 81
    wb["ma_0009"].append(["O50147_lysZ", 82.0, "medium"])
    wb.save(path)


def test_species_datasets(tmp_path):
    p = tmp_path / "s1.xlsx"
    _make_s1(p)
    spec = species_datasets(p)
    assert spec["Thermus thermophilus"]["ogt"] == 70.0
    assert sorted(spec["Thermus thermophilus"]["dataset_ids"]) == ["ma_0008", "ma_0009"]
    assert spec["Thermus thermophilus"]["class"] == "thermophile"
    assert spec["Escherichia coli"]["class"] == "mesophile"
    assert spec["Oleispira antarctica"]["class"] == "psychrophile"


def test_organism_tms_medians_across_datasets(tmp_path):
    p = tmp_path / "s4.xlsx"
    _make_s4(p)
    tms = organism_tms(p, ["ma_0008", "ma_0009"])
    assert tms["O50147"] == 81.0      # median of 80 and 82
    assert "O66271" not in tms        # non-melter (None Tm) dropped
