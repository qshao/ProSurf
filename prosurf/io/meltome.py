import re
import statistics
from pathlib import Path

import openpyxl

_CANONICAL_RE = re.compile(
    r"^([A-NR-Z][0-9][A-Z][A-Z0-9]{2}[0-9]|[OPQ][0-9][A-Z0-9]{3}[0-9])$"
)


def _organism_class(ogt: float) -> str:
    if ogt < 20:
        return "psychrophile"
    if ogt <= 45:
        return "mesophile"
    return "thermophile"


def canonical_uniprot(protein_id: str) -> str | None:
    """Extract canonical UniProt accession from a Meltome protein ID such as
    'P61626_LYZ' or 'Q9UNM6-2_PSMD13'. Returns None if not canonical."""
    uid = protein_id.split("_")[0].split("-")[0]
    return uid if _CANONICAL_RE.match(uid) else None


def species_datasets(s1_path: Path) -> dict:
    """Parse the S1 'Meltome data set' sheet into
    {organism: {'ogt': float, 'dataset_ids': [str], 'class': str}}.
    Only the primary 'Meltome data set' sheet is read; reproducibility and
    inter-lab sheets are ignored."""
    wb = openpyxl.load_workbook(s1_path, read_only=True)
    ws = wb["Meltome data set"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    if header[0] != "Dataset ID" or "OGT [°C]" not in header:
        raise ValueError(f"Unexpected S1 header: {header[:5]}")
    org_idx = header.index("Organism")
    ogt_idx = header.index("OGT [°C]")
    out: dict = {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        did = str(r[0]).replace("ma:", "ma_", 1)
        org = str(r[org_idx]).strip()
        ogt = r[ogt_idx]
        if not org or not isinstance(ogt, (int, float)):
            continue
        d = out.setdefault(
            org,
            {"ogt": float(ogt), "dataset_ids": [], "class": _organism_class(float(ogt))},
        )
        d["dataset_ids"].append(did)
    return out


def organism_tms(s4_path: Path, dataset_ids: list) -> dict:
    """Median Tm per canonical UniProt across the given S4 dataset sheets.
    Sheet layout: column 0 = Protein ID, column 1 = Melting point [°C]
    (None for non-melters, which are dropped)."""
    wb = openpyxl.load_workbook(s4_path, read_only=True)
    per_protein: dict = {}
    for did in dataset_ids:
        if did not in wb.sheetnames:
            continue
        ws = wb[did]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            uid = canonical_uniprot(str(row[0]))
            tm = row[1] if len(row) > 1 else None
            if uid is None or not isinstance(tm, (int, float)):
                continue
            per_protein.setdefault(uid, []).append(float(tm))
    return {uid: statistics.median(v) for uid, v in per_protein.items()}
